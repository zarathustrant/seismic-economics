from flask import Flask, send_file, request, jsonify
from pymongo import MongoClient, ASCENDING, DESCENDING, UpdateOne
from pymongo.errors import PyMongoError
from dotenv import load_dotenv
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
import os
import re
import json
import time
import logging

app = Flask(__name__)
app.logger.setLevel(logging.INFO)

# Password (set via environment variable or default)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
HTML_FILE = os.path.join(BASE_DIR, 'app.html')
load_dotenv(os.path.join(BASE_DIR, '.env'))

MONGODB_URI = os.environ.get('MONGODB_URI', '').strip()
MONGO_DB_NAME = os.environ.get('MONGODB_DB', 'seismic_economics')
MONGO_COLLECTION = os.environ.get('MONGODB_COLLECTION', 'projects')

mongo_client = None
projects_col = None
indexes_ready = False
migration_checked = False


def now_ms():
    return int(time.time() * 1000)


def generate_project_id():
    return f"p{now_ms():x}"


def clean_name(value):
    text = str(value or '').strip()
    return text[:120] if text else 'Untitled Project'


def clean_project_id(value):
    text = str(value or '').strip()
    if not text:
        return generate_project_id()
    safe = ''.join(ch for ch in text if ch.isalnum() or ch in ('-', '_'))
    return safe[:60] or generate_project_id()


def parse_embedded_projects():
    if not os.path.exists(HTML_FILE):
        return []
    with open(HTML_FILE, 'r', encoding='utf-8') as f:
        html = f.read()

    match = re.search(
        r'<script type="application/json" id="embedded-projects-data">([\s\S]*?)</script>',
        html,
        re.IGNORECASE,
    )
    if not match:
        return []

    raw = match.group(1).strip()
    if not raw:
        return []

    payload = json.loads(raw)
    projects = payload.get('projects') if isinstance(payload, dict) else {}
    if not isinstance(projects, dict):
        return []

    items = []
    for key, proj in projects.items():
        if not isinstance(proj, dict):
            continue
        project_data = proj.get('data') if isinstance(proj.get('data'), dict) else {}
        pid = clean_project_id(proj.get('id') or key)
        name = clean_name(proj.get('name') or project_data.get('projectInputs', {}).get('projectName'))
        created_at = int(proj.get('createdAt') or now_ms())
        updated_at = int(proj.get('updatedAt') or created_at)
        items.append({
            '_id': pid,
            'name': name,
            'data': project_data,
            'createdAt': created_at,
            'updatedAt': updated_at,
        })

    return items


def ensure_mongo_handles():
    global mongo_client, projects_col

    if projects_col is not None:
        return projects_col
    if not MONGODB_URI:
        return None

    mongo_client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=5000)
    projects_col = mongo_client[MONGO_DB_NAME][MONGO_COLLECTION]
    return projects_col


def ensure_db_ready():
    global indexes_ready, migration_checked

    col = ensure_mongo_handles()
    if col is None:
        return False, 'MONGODB_URI is not configured on the server.'

    try:
        if not indexes_ready:
            col.create_index([('updatedAt', DESCENDING)])
            col.create_index([('createdAt', ASCENDING)])
            indexes_ready = True

        if not migration_checked:
            has_any = col.count_documents({}, limit=1) > 0
            if not has_any:
                seed = parse_embedded_projects()
                if seed:
                    col.insert_many(seed, ordered=False)
            migration_checked = True

        return True, None
    except PyMongoError as exc:
        return False, str(exc)


def sanitize_project_payload(payload, *, require_data=False):
    if not isinstance(payload, dict):
        return None, 'Payload must be a JSON object.'

    name = clean_name(payload.get('name')) if ('name' in payload or require_data) else None
    project_data = payload.get('data')

    if require_data and not isinstance(project_data, dict):
        return None, 'Project data must be an object.'
    if project_data is not None and not isinstance(project_data, dict):
        return None, 'Project data must be an object.'

    out = {
        'name': name,
        'data': project_data,
    }
    return out, None


def serialize_project_meta(doc):
    return {
        'id': doc['_id'],
        'name': doc.get('name', 'Untitled Project'),
        'createdAt': int(doc.get('createdAt') or now_ms()),
        'updatedAt': int(doc.get('updatedAt') or now_ms()),
    }


def serialize_project_full(doc):
    item = serialize_project_meta(doc)
    item['data'] = doc.get('data') if isinstance(doc.get('data'), dict) else {}
    return item


def summarize_project_data(data):
    if not isinstance(data, dict):
        return {'valid': False}
    rate_cards = data.get('rateCards') if isinstance(data.get('rateCards'), dict) else {}
    cost_lines = data.get('costLines') if isinstance(data.get('costLines'), list) else []

    nonzero_rates = 0
    for arr in rate_cards.values():
        if not isinstance(arr, list):
            continue
        for rate in arr:
            if not isinstance(rate, dict):
                continue
            try:
                if float(rate.get('value') or 0) != 0:
                    nonzero_rates += 1
            except Exception:
                pass

    nonzero_month_cells = 0
    for line in cost_lines:
        if not isinstance(line, dict):
            continue
        for v in (line.get('monthly') or []):
            try:
                if float(v or 0) != 0:
                    nonzero_month_cells += 1
            except Exception:
                pass

    return {
        'valid': True,
        'projectName': data.get('projectInputs', {}).get('projectName'),
        'nonzeroRates': nonzero_rates,
        'nonzeroMonthlyCells': nonzero_month_cells,
        'costLines': len(cost_lines),
    }


def to_num(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def clamp_int(value, minimum, maximum):
    num = int(round(to_num(value, minimum)))
    return max(minimum, min(maximum, num))


def month_labels(duration):
    d = clamp_int(duration, 1, 14)
    return [f"M{i:02d}" for i in range(1, d + 1)] + ["Demob"], d


def profile_weight(profile, pos, active, ramp):
    r = max(1, ramp)
    if profile == 'Uniform':
        return 1
    if profile == 'Manual':
        return 1
    if profile == 'Front-Loaded':
        return (active - pos + 1) / active
    if profile == 'Back-Loaded':
        return pos / active
    if profile == 'Bell Curve':
        return max(0.05, 1 - (2 * abs(pos - ((active + 1) / 2)) / active))
    if pos <= r:
        return pos / r
    if pos > active - r:
        return (active - pos + 1) / r
    return 1


def phased_series(program, months):
    out = [0.0] * months
    if not isinstance(program, dict):
        return out

    profile = str(program.get('profile') or 'S-Curve')
    if profile == 'Manual':
        manual = program.get('manualMonthly') if isinstance(program.get('manualMonthly'), list) else []
        for i in range(min(months, len(manual))):
            out[i] = to_num(manual[i], 0.0)
        return out

    total = to_num(program.get('total'), 0.0)
    start = clamp_int(program.get('startMonth'), 1, months)
    active = clamp_int(program.get('activeMonths'), 1, months)
    ramp = clamp_int(program.get('rampMonths'), 1, active)

    active_slots = []
    weights = []
    for m in range(1, months + 1):
        if m >= start and m < start + active and m <= months - 1:
            pos = m - start + 1
            active_slots.append(m - 1)
            weights.append(profile_weight(profile, pos, active, ramp))

    total_weight = sum(weights) or 1
    allocated = 0.0
    for i, idx in enumerate(active_slots):
        is_last = i == len(active_slots) - 1
        value = max(0, total - allocated) if is_last else round(total * weights[i] / total_weight)
        out[idx] = float(value)
        allocated += value

    return out


def normalized_series(values, months):
    out = [0.0] * months
    if not isinstance(values, list):
        return out
    for i in range(min(months, len(values))):
        out[i] = to_num(values[i], 0.0)
    return out


def build_production(data, months):
    programs = data.get('productionPrograms') if isinstance(data.get('productionPrograms'), dict) else {}
    inputs = data.get('projectInputs') if isinstance(data.get('projectInputs'), dict) else {}

    survey = phased_series(programs.get('survey', {}), months)
    drilling = phased_series(programs.get('drilling', {}), months)
    recording = phased_series(programs.get('recording', {}), months)
    upholes = phased_series(programs.get('upholes', {}), months)
    shot_factor = to_num(inputs.get('shotFactor'), 0)
    shots = [v / shot_factor if shot_factor else 0.0 for v in recording]
    return {
        'survey': survey,
        'drilling': drilling,
        'recording': recording,
        'shots': shots,
        'upholes': upholes,
    }


def line_share_factor(line, cost_share):
    mode = str((line or {}).get('shareMode') or '').lower()
    return cost_share if mode == 'shared' else 1.0


def driver_series(line, prod, months):
    ref = str((line or {}).get('driverRef') or '')
    if ref == 'surveyKm':
        return list(prod['survey'])
    if ref == 'drillSdh':
        return list(prod['drilling'])
    if ref == 'recordingKm':
        return list(prod['recording'])
    if ref == 'recordingShots':
        return list(prod['shots'])
    if ref == 'upholes':
        return list(prod['upholes'])
    return [0.0] * months


def compute_line_series(line, prod, months, cost_share):
    driver_type = str((line or {}).get('driverType') or '')
    rate = to_num((line or {}).get('rate'), 0.0)
    monthly = normalized_series((line or {}).get('monthly'), months)

    if driver_type in ('productionLinked', 'crewLinked'):
        base = [v * rate for v in driver_series(line, prod, months)]
    elif driver_type == 'oneTime' and re.match(r'^m\d+$', str((line or {}).get('driverRef') or '')):
        base = [0.0] * months
        idx = int(str(line.get('driverRef'))[1:]) - 1
        if 0 <= idx < months:
            base[idx] = rate if rate != 0 else monthly[idx]
    elif driver_type == 'fixedMonthly' and rate > 0:
        has_manual = any(to_num(v, 0) != 0 for v in ((line or {}).get('monthly') or []))
        base = monthly if has_manual else [rate] * months
    else:
        base = monthly

    factor = line_share_factor(line, cost_share)
    return [v * factor for v in base]


def calculate_model(data):
    inputs = data.get('projectInputs') if isinstance(data.get('projectInputs'), dict) else {}
    terms = data.get('contractTerms') if isinstance(data.get('contractTerms'), dict) else {}
    cost_lines = data.get('costLines') if isinstance(data.get('costLines'), list) else []

    labels, duration = month_labels(inputs.get('duration'))
    months = len(labels)
    prod = build_production(data, months)
    cost_share = to_num(inputs.get('costShare'), 0.0)
    revenue_share = to_num(inputs.get('revenueShare'), 0.0)
    fx = to_num(inputs.get('fx'), 0.0)

    line_series = []
    category_series = {}
    categories = []
    for line in cost_lines:
        if not isinstance(line, dict):
            continue
        cat = str(line.get('category') or 'Uncategorized')
        if cat not in category_series:
            category_series[cat] = [0.0] * months
            categories.append(cat)
        vals = compute_line_series(line, prod, months, cost_share)
        line_series.append(vals)
        for i, v in enumerate(vals):
            category_series[cat][i] += v

    funding = [0.0] * months
    for cat in categories:
        for i, v in enumerate(category_series[cat]):
            funding[i] += v

    turnkey = to_num(terms.get('turnkeyRate'), 0.0)
    mob = to_num(terms.get('mobilisation'), 0.0)
    uphole_rate = to_num(terms.get('upholeRate'), 0.0)
    demob = to_num(terms.get('demobilisation'), 0.0)

    revenue_prod_usd = [v * turnkey for v in prod['recording']]
    mob_usd = [0.0] * months
    mob_usd[0] = mob if months else 0.0
    uphole_usd = [v * uphole_rate for v in prod['upholes']]
    demob_usd = [0.0] * months
    demob_idx = max(0, min(months - 1, duration - 1))
    demob_usd[demob_idx] = demob if months else 0.0

    revenue_100_usd = [revenue_prod_usd[i] + mob_usd[i] + uphole_usd[i] + demob_usd[i] for i in range(months)]
    revenue_share_usd = [v * revenue_share for v in revenue_100_usd]
    revenue_ngn = [v * fx for v in revenue_share_usd]
    cash = [revenue_ngn[i] - funding[i] for i in range(months)]

    return {
        'labels': labels,
        'months': months,
        'duration': duration,
        'inputs': inputs,
        'terms': terms,
        'prod': prod,
        'costLines': cost_lines,
        'lineSeries': line_series,
        'categories': categories,
        'categorySeries': category_series,
        'funding': funding,
        'revenueNgn': revenue_ngn,
        'cash': cash,
    }


def style_header_row(ws, row, start_col, end_col):
    fill = PatternFill(start_color='155E75', end_color='155E75', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE'),
    )
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(color='FFFFFF', bold=True, size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def autofit(ws, extra=2, max_width=48):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            val = '' if c.value is None else str(c.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + extra))


def safe_filename(value):
    cleaned = re.sub(r'[^A-Za-z0-9._-]+', '_', str(value or 'project')).strip('._')
    return cleaned[:80] or 'project'


def build_excel_workbook(project_name, data):
    calc = calculate_model(data if isinstance(data, dict) else {})
    labels = calc['labels']
    months = calc['months']
    inputs = calc['inputs']
    terms = calc['terms']
    prod = calc['prod']
    cost_lines = calc['costLines']

    wb = Workbook()
    ws_over = wb.active
    ws_over.title = 'Overview'
    ws_inputs = wb.create_sheet('Inputs')
    ws_rates = wb.create_sheet('Rates')
    ws_prod = wb.create_sheet('Production')
    ws_cost = wb.create_sheet('Cost Model')
    ws_rf = wb.create_sheet('Revenue & Funding')

    money_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )

    # Inputs sheet
    ws_inputs['A1'] = 'Project Inputs & Commercial Terms'
    ws_inputs['A1'].font = Font(size=14, bold=True, color='1F2937')
    ws_inputs['A2'] = 'Edit inputs here to stress-test assumptions.'
    ws_inputs['A2'].font = Font(size=10, color='64748B')
    ws_inputs.freeze_panes = 'A4'
    style_header_row(ws_inputs, 3, 1, 3)
    ws_inputs['A3'] = 'Field'
    ws_inputs['B3'] = 'Value'
    ws_inputs['C3'] = 'Unit'

    input_rows = [
        ('projectName', 'Project Name', str(inputs.get('projectName') or project_name), ''),
        ('projectType', 'Project Type', str(inputs.get('projectType') or '2D').upper(), ''),
        ('omlNumber', 'OML / Basin', str(inputs.get('omlNumber') or ''), ''),
        ('client', 'Client', str(inputs.get('client') or ''), ''),
        ('contractor', 'Contractor', str(inputs.get('contractor') or ''), ''),
        ('projectArea', 'Project Area', to_num(inputs.get('projectArea'), 0.0), 'Km2' if str(inputs.get('projectType') or '').upper() == '3D' else 'Km'),
        ('duration', 'Duration', calc['duration'], 'months'),
        ('shotFactor', 'Shot Factor', to_num(inputs.get('shotFactor'), 0.0), 'km/shot'),
        ('revenueShare', 'Revenue Share', to_num(inputs.get('revenueShare'), 0.0), 'x'),
        ('costShare', 'Cost Share', to_num(inputs.get('costShare'), 0.0), 'x'),
        ('fx', 'FX Rate', to_num(inputs.get('fx'), 0.0), 'NGN/USD'),
        ('turnkeyRate', 'Turnkey Rate', to_num(terms.get('turnkeyRate'), 0.0), 'USD/Km'),
        ('mobilisation', 'Mobilisation', to_num(terms.get('mobilisation'), 0.0), 'USD'),
        ('demobilisation', 'Demobilisation', to_num(terms.get('demobilisation'), 0.0), 'USD'),
        ('upholeRate', 'Uphole Rate', to_num(terms.get('upholeRate'), 0.0), 'USD/uphole'),
    ]

    input_row_map = {}
    row = 4
    for key, label, value, unit in input_rows:
        ws_inputs.cell(row=row, column=1, value=label)
        ws_inputs.cell(row=row, column=2, value=value)
        ws_inputs.cell(row=row, column=3, value=unit)
        ws_inputs.cell(row=row, column=1).border = thin
        ws_inputs.cell(row=row, column=2).border = thin
        ws_inputs.cell(row=row, column=3).border = thin
        if isinstance(value, (float, int)):
            ws_inputs.cell(row=row, column=2).number_format = money_fmt if key in ('turnkeyRate', 'mobilisation', 'demobilisation', 'upholeRate', 'fx') else '#,##0.0000'
        input_row_map[key] = row
        row += 1

    # Rates sheet
    ws_rates['A1'] = 'Rates Dictionary'
    ws_rates['A1'].font = Font(size=14, bold=True, color='1F2937')
    style_header_row(ws_rates, 3, 1, 4)
    ws_rates['A3'] = 'Section'
    ws_rates['B3'] = 'Rate Item'
    ws_rates['C3'] = 'Value'
    ws_rates['D3'] = 'Unit'

    rate_cards = data.get('rateCards') if isinstance(data.get('rateCards'), dict) else {}
    row = 4
    for section, arr in rate_cards.items():
        if not isinstance(arr, list):
            continue
        for rate in arr:
            if not isinstance(rate, dict):
                continue
            ws_rates.cell(row=row, column=1, value=str(section))
            ws_rates.cell(row=row, column=2, value=str(rate.get('name') or ''))
            ws_rates.cell(row=row, column=3, value=to_num(rate.get('value'), 0.0))
            ws_rates.cell(row=row, column=4, value=str(rate.get('unit') or ''))
            ws_rates.cell(row=row, column=3).number_format = money_fmt
            for c in range(1, 5):
                ws_rates.cell(row=row, column=c).border = thin
            row += 1

    # Production sheet
    ws_prod['A1'] = 'Production Program'
    ws_prod['A1'].font = Font(size=14, bold=True, color='1F2937')
    ws_prod['A2'] = 'Monthly production profile used in revenue and production-linked costs.'
    ws_prod['A2'].font = Font(size=10, color='64748B')
    start_col = 2
    for i, m in enumerate(labels):
        ws_prod.cell(row=3, column=start_col + i, value=m)
    total_col = start_col + months
    ws_prod.cell(row=3, column=1, value='Metric')
    ws_prod.cell(row=3, column=total_col, value='Total')
    style_header_row(ws_prod, 3, 1, total_col)

    prod_rows = {
        'survey': 4,
        'drilling': 5,
        'recording': 6,
        'shots': 7,
        'upholes': 8,
    }
    prod_labels = {
        'survey': 'Survey (Km)',
        'drilling': 'Drilling (SDH)',
        'recording': 'Recording (Km)',
        'shots': 'Recording Shots',
        'upholes': 'Upholes',
    }
    for key, r in prod_rows.items():
        ws_prod.cell(row=r, column=1, value=prod_labels[key])
        vals = prod[key]
        for i, v in enumerate(vals):
            ws_prod.cell(row=r, column=start_col + i, value=v)
        ws_prod.cell(row=r, column=total_col, value=f"=SUM({get_column_letter(start_col)}{r}:{get_column_letter(total_col - 1)}{r})")
        for c in range(1, total_col + 1):
            ws_prod.cell(row=r, column=c).border = thin
            if c > 1:
                ws_prod.cell(row=r, column=c).number_format = '#,##0.000'

    # Cost model sheet
    ws_cost['A1'] = 'Cost Model (Adjusted by Share Mode)'
    ws_cost['A1'].font = Font(size=14, bold=True, color='1F2937')
    ws_cost['A2'] = 'All line values below are post-share adjusted. Total column is formula-driven.'
    ws_cost['A2'].font = Font(size=10, color='64748B')

    month_col_start = 6
    ws_cost.cell(row=4, column=1, value='Category')
    ws_cost.cell(row=4, column=2, value='Line Item')
    ws_cost.cell(row=4, column=3, value='Badge')
    ws_cost.cell(row=4, column=4, value='Share')
    ws_cost.cell(row=4, column=5, value='Driver')
    for i, m in enumerate(labels):
        ws_cost.cell(row=4, column=month_col_start + i, value=m)
    cost_total_col = month_col_start + months
    ws_cost.cell(row=4, column=cost_total_col, value='Total')
    style_header_row(ws_cost, 4, 1, cost_total_col)

    line_start_row = 5
    row = line_start_row
    for idx, line in enumerate(cost_lines):
        if not isinstance(line, dict):
            continue
        vals = calc['lineSeries'][idx] if idx < len(calc['lineSeries']) else [0.0] * months
        ws_cost.cell(row=row, column=1, value=str(line.get('category') or 'Uncategorized'))
        ws_cost.cell(row=row, column=2, value=str(line.get('name') or ''))
        ws_cost.cell(row=row, column=3, value=str(line.get('badge') or '').upper())
        ws_cost.cell(row=row, column=4, value='Shared' if str(line.get('shareMode') or '').lower() == 'shared' else 'Full')
        ws_cost.cell(row=row, column=5, value=str(line.get('driverType') or 'manual'))
        for i, v in enumerate(vals):
            ws_cost.cell(row=row, column=month_col_start + i, value=v)
        ws_cost.cell(
            row=row,
            column=cost_total_col,
            value=f"=SUM({get_column_letter(month_col_start)}{row}:{get_column_letter(cost_total_col - 1)}{row})"
        )
        for c in range(1, cost_total_col + 1):
            ws_cost.cell(row=row, column=c).border = thin
            if c >= month_col_start:
                ws_cost.cell(row=row, column=c).number_format = money_fmt
        row += 1
    line_end_row = max(line_start_row, row - 1)

    summary_start = row + 2
    ws_cost.cell(row=summary_start, column=1, value='Category Summary').font = Font(bold=True, size=11, color='1F2937')
    ws_cost.cell(row=summary_start + 1, column=1, value='Category')
    for i, m in enumerate(labels):
        ws_cost.cell(row=summary_start + 1, column=month_col_start + i, value=m)
    ws_cost.cell(row=summary_start + 1, column=cost_total_col, value='Total')
    style_header_row(ws_cost, summary_start + 1, 1, cost_total_col)

    cat_rows_start = summary_start + 2
    categories = calc['categories']
    for i, cat in enumerate(categories):
        r = cat_rows_start + i
        ws_cost.cell(row=r, column=1, value=cat)
        for m_idx in range(months):
            col = get_column_letter(month_col_start + m_idx)
            ws_cost.cell(
                row=r,
                column=month_col_start + m_idx,
                value=f'=SUMIF($A${line_start_row}:$A${line_end_row},$A{r},{col}${line_start_row}:{col}${line_end_row})'
            )
            ws_cost.cell(row=r, column=month_col_start + m_idx).number_format = money_fmt
        tot_col_letter = get_column_letter(cost_total_col)
        ws_cost.cell(
            row=r,
            column=cost_total_col,
            value=f'=SUMIF($A${line_start_row}:$A${line_end_row},$A{r},{tot_col_letter}${line_start_row}:{tot_col_letter}${line_end_row})'
        )
        ws_cost.cell(row=r, column=cost_total_col).number_format = money_fmt
        for c in range(1, cost_total_col + 1):
            ws_cost.cell(row=r, column=c).border = thin

    grand_row = cat_rows_start + len(categories)
    ws_cost.cell(row=grand_row, column=1, value='GRAND TOTAL').font = Font(bold=True, color='1F2937')
    for m_idx in range(months):
        col = get_column_letter(month_col_start + m_idx)
        ws_cost.cell(
            row=grand_row,
            column=month_col_start + m_idx,
            value=f"=SUM({col}{cat_rows_start}:{col}{grand_row - 1})"
        )
        ws_cost.cell(row=grand_row, column=month_col_start + m_idx).number_format = money_fmt
    ws_cost.cell(
        row=grand_row,
        column=cost_total_col,
        value=f"=SUM({get_column_letter(cost_total_col)}{cat_rows_start}:{get_column_letter(cost_total_col)}{grand_row - 1})"
    )
    ws_cost.cell(row=grand_row, column=cost_total_col).number_format = money_fmt
    for c in range(1, cost_total_col + 1):
        ws_cost.cell(row=grand_row, column=c).border = thin
        ws_cost.cell(row=grand_row, column=c).fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

    # Revenue and funding sheet
    ws_rf['A1'] = 'Revenue, Funding, and Cash Flow Build'
    ws_rf['A1'].font = Font(size=14, bold=True, color='1F2937')
    ws_rf['A2'] = 'Formulas are linked across Inputs, Production, and Cost Model sheets.'
    ws_rf['A2'].font = Font(size=10, color='64748B')
    ws_rf.cell(row=3, column=1, value='Metric')
    for i, m in enumerate(labels):
        ws_rf.cell(row=3, column=2 + i, value=m)
    rf_total_col = 2 + months
    ws_rf.cell(row=3, column=rf_total_col, value='Total')
    style_header_row(ws_rf, 3, 1, rf_total_col)

    rf_rows = {
        'survey': 4,
        'recording': 5,
        'upholes': 6,
        'rev_prod': 8,
        'mob': 9,
        'uphole_rev': 10,
        'demob': 11,
        'rev100': 12,
        'rev_share': 13,
        'rev_ngn': 14,
        'funding': 16,
        'net': 17,
        'cum': 18,
    }

    ws_rf.cell(row=rf_rows['survey'], column=1, value='Survey (Km)')
    ws_rf.cell(row=rf_rows['recording'], column=1, value='Recording (Km)')
    ws_rf.cell(row=rf_rows['upholes'], column=1, value='Upholes')
    ws_rf.cell(row=rf_rows['rev_prod'], column=1, value='Production Revenue (USD)')
    ws_rf.cell(row=rf_rows['mob'], column=1, value='Mobilisation (USD)')
    ws_rf.cell(row=rf_rows['uphole_rev'], column=1, value='Uphole Revenue (USD)')
    ws_rf.cell(row=rf_rows['demob'], column=1, value='Demobilisation (USD)')
    ws_rf.cell(row=rf_rows['rev100'], column=1, value='Total Revenue 100% (USD)')
    ws_rf.cell(row=rf_rows['rev_share'], column=1, value='Revenue Share (USD)')
    ws_rf.cell(row=rf_rows['rev_ngn'], column=1, value='Revenue (NGN)')
    ws_rf.cell(row=rf_rows['funding'], column=1, value='Funding Requirement (NGN)')
    ws_rf.cell(row=rf_rows['net'], column=1, value='Net Cash Flow (NGN)')
    ws_rf.cell(row=rf_rows['cum'], column=1, value='Cumulative Cash (NGN)')

    for idx in range(months):
        col = get_column_letter(2 + idx)
        prod_col = get_column_letter(2 + idx)
        ws_rf.cell(row=rf_rows['survey'], column=2 + idx, value=f"='Production'!{prod_col}{prod_rows['survey']}")
        ws_rf.cell(row=rf_rows['recording'], column=2 + idx, value=f"='Production'!{prod_col}{prod_rows['recording']}")
        ws_rf.cell(row=rf_rows['upholes'], column=2 + idx, value=f"='Production'!{prod_col}{prod_rows['upholes']}")
        ws_rf.cell(
            row=rf_rows['rev_prod'],
            column=2 + idx,
            value=f"={col}{rf_rows['recording']}*'Inputs'!$B${input_row_map['turnkeyRate']}"
        )
        ws_rf.cell(
            row=rf_rows['mob'],
            column=2 + idx,
            value=f"=IF({idx + 1}=1,'Inputs'!$B${input_row_map['mobilisation']},0)"
        )
        ws_rf.cell(
            row=rf_rows['uphole_rev'],
            column=2 + idx,
            value=f"={col}{rf_rows['upholes']}*'Inputs'!$B${input_row_map['upholeRate']}"
        )
        ws_rf.cell(
            row=rf_rows['demob'],
            column=2 + idx,
            value=f"=IF({idx + 1}='Inputs'!$B${input_row_map['duration']},'Inputs'!$B${input_row_map['demobilisation']},0)"
        )
        ws_rf.cell(
            row=rf_rows['rev100'],
            column=2 + idx,
            value=f"=SUM({col}{rf_rows['rev_prod']}:{col}{rf_rows['demob']})"
        )
        ws_rf.cell(
            row=rf_rows['rev_share'],
            column=2 + idx,
            value=f"={col}{rf_rows['rev100']}*'Inputs'!$B${input_row_map['revenueShare']}"
        )
        ws_rf.cell(
            row=rf_rows['rev_ngn'],
            column=2 + idx,
            value=f"={col}{rf_rows['rev_share']}*'Inputs'!$B${input_row_map['fx']}"
        )
        cost_col = get_column_letter(month_col_start + idx)
        ws_rf.cell(
            row=rf_rows['funding'],
            column=2 + idx,
            value=f"='Cost Model'!{cost_col}{grand_row}"
        )
        ws_rf.cell(
            row=rf_rows['net'],
            column=2 + idx,
            value=f"={col}{rf_rows['rev_ngn']}-{col}{rf_rows['funding']}"
        )
        if idx == 0:
            ws_rf.cell(row=rf_rows['cum'], column=2 + idx, value=f"={col}{rf_rows['net']}")
        else:
            prev_col = get_column_letter(1 + idx)
            ws_rf.cell(row=rf_rows['cum'], column=2 + idx, value=f"={prev_col}{rf_rows['cum']}+{col}{rf_rows['net']}")

    for key, r in rf_rows.items():
        for idx in range(months):
            c = 2 + idx
            ws_rf.cell(row=r, column=c).border = thin
            if key in ('survey', 'recording', 'upholes'):
                ws_rf.cell(row=r, column=c).number_format = '#,##0.000'
            else:
                ws_rf.cell(row=r, column=c).number_format = money_fmt

        ws_rf.cell(
            row=r,
            column=rf_total_col,
            value=f"=SUM(B{r}:{get_column_letter(rf_total_col - 1)}{r})"
        )
        ws_rf.cell(row=r, column=rf_total_col).border = thin
        if key in ('survey', 'recording', 'upholes'):
            ws_rf.cell(row=r, column=rf_total_col).number_format = '#,##0.000'
        else:
            ws_rf.cell(row=r, column=rf_total_col).number_format = money_fmt

    ws_rf.cell(
        row=rf_rows['cum'],
        column=rf_total_col,
        value=f"={get_column_letter(rf_total_col - 1)}{rf_rows['cum']}"
    )
    ws_rf.cell(row=rf_rows['cum'], column=rf_total_col).number_format = money_fmt

    # Overview sheet
    ws_over['A1'] = 'Seismic Project Economics Model'
    ws_over['A1'].font = Font(size=16, bold=True, color='1F2937')
    ws_over['A2'] = f"Exported: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"
    ws_over['A2'].font = Font(size=10, color='64748B')
    ws_over['A4'] = 'Project'
    ws_over['B4'] = str(inputs.get('projectName') or project_name)
    ws_over['A5'] = 'Project Type'
    ws_over['B5'] = str(inputs.get('projectType') or '2D').upper()
    ws_over['A6'] = 'Duration'
    ws_over['B6'] = calc['duration']
    ws_over['A7'] = 'Months in Model'
    ws_over['B7'] = months
    ws_over['A4'].font = ws_over['A5'].font = ws_over['A6'].font = ws_over['A7'].font = Font(bold=True, color='334155')

    ws_over['D4'] = 'KPI'
    ws_over['E4'] = 'Value'
    style_header_row(ws_over, 4, 4, 5)
    ws_over['D5'] = 'Total Revenue (NGN)'
    ws_over['E5'] = f"=SUM('Revenue & Funding'!B{rf_rows['rev_ngn']}:{get_column_letter(rf_total_col - 1)}{rf_rows['rev_ngn']})"
    ws_over['D6'] = 'Funding Requirement (NGN)'
    ws_over['E6'] = f"=SUM('Revenue & Funding'!B{rf_rows['funding']}:{get_column_letter(rf_total_col - 1)}{rf_rows['funding']})"
    ws_over['D7'] = 'Net Cash Flow (NGN)'
    ws_over['E7'] = "=E5-E6"
    ws_over['D8'] = 'CRR'
    ws_over['E8'] = "=IF(E5=0,0,E6/E5)"
    ws_over['D9'] = 'Cost Lines'
    ws_over['E9'] = len(cost_lines)
    ws_over['D10'] = 'Categories'
    ws_over['E10'] = len(categories)
    for r in range(5, 11):
        ws_over.cell(row=r, column=4).border = thin
        ws_over.cell(row=r, column=5).border = thin
    ws_over['E5'].number_format = ws_over['E6'].number_format = ws_over['E7'].number_format = money_fmt
    ws_over['E8'].number_format = pct_fmt

    badge_label = {
        'fixed': 'Fixed Cost',
        'var': 'Variable Cost',
        'sub': 'Subcontract',
        'lease': 'Lease',
        'owned': 'Owned Asset',
        'labor': 'Labour',
        'unassigned': 'Unassigned',
    }
    ws_over['A12'] = 'Cost Distribution by Badge (Formula-Based)'
    ws_over['A12'].font = Font(bold=True, color='1F2937')
    ws_over['A13'] = 'Badge'
    ws_over['B13'] = 'Amount (NGN)'
    ws_over['C13'] = 'Share %'
    style_header_row(ws_over, 13, 1, 3)

    badges = sorted({str((line or {}).get('badge') or 'unassigned').lower() for line in cost_lines}) or ['unassigned']
    badge_start = 14
    total_col_letter = get_column_letter(cost_total_col)
    badge_total_row = badge_start + len(badges)
    for i, badge in enumerate(badges):
        r = badge_start + i
        key = badge.upper()
        ws_over.cell(row=r, column=1, value=badge_label.get(badge, key.title()))
        ws_over.cell(
            row=r,
            column=2,
            value=f'=SUMIF(\'Cost Model\'!$C${line_start_row}:$C${line_end_row},"{key}",\'Cost Model\'!${total_col_letter}${line_start_row}:${total_col_letter}${line_end_row})'
        )
        ws_over.cell(row=r, column=3, value=f"=IF(B{badge_total_row}=0,0,B{r}/B{badge_total_row})")
        ws_over.cell(row=r, column=2).number_format = money_fmt
        ws_over.cell(row=r, column=3).number_format = pct_fmt
        for c in range(1, 4):
            ws_over.cell(row=r, column=c).border = thin
    ws_over.cell(row=badge_total_row, column=1, value='Total').font = Font(bold=True)
    ws_over.cell(row=badge_total_row, column=2, value=f"=SUM(B{badge_start}:B{badge_total_row - 1})")
    ws_over.cell(row=badge_total_row, column=3, value='=1')
    ws_over.cell(row=badge_total_row, column=2).number_format = money_fmt
    ws_over.cell(row=badge_total_row, column=3).number_format = pct_fmt
    for c in range(1, 4):
        ws_over.cell(row=badge_total_row, column=c).border = thin
        ws_over.cell(row=badge_total_row, column=c).fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

    pie = PieChart()
    pie.title = 'Badge Distribution'
    pie.height = 7
    pie.width = 9
    pie_data = Reference(ws_over, min_col=2, min_row=badge_start, max_row=badge_total_row - 1)
    pie_labels = Reference(ws_over, min_col=1, min_row=badge_start, max_row=badge_total_row - 1)
    pie.add_data(pie_data, titles_from_data=False)
    pie.set_categories(pie_labels)
    ws_over.add_chart(pie, 'E12')

    for ws in (ws_over, ws_inputs, ws_rates, ws_prod, ws_cost, ws_rf):
        autofit(ws)

    ws_over.freeze_panes = 'A4'
    ws_inputs.freeze_panes = 'A4'
    ws_rates.freeze_panes = 'A4'
    ws_prod.freeze_panes = 'B4'
    ws_cost.freeze_panes = 'F5'
    ws_rf.freeze_panes = 'B4'

    return wb


@app.route('/')
def index():
    """Serve the main HTML file"""
    return send_file(HTML_FILE)


@app.route('/api/projects', methods=['GET'])
def list_projects():
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    docs = list(projects_col.find({}, {'data': 0}).sort('updatedAt', DESCENDING))
    app.logger.info("GET /api/projects -> %s projects", len(docs))
    return jsonify({'success': True, 'projects': [serialize_project_meta(d) for d in docs]})


@app.route('/api/projects/<project_id>', methods=['GET'])
def get_project(project_id):
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    pid = clean_project_id(project_id)
    doc = projects_col.find_one({'_id': pid})
    if not doc:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    app.logger.info("GET /api/projects/%s -> found=%s", pid, bool(doc))
    return jsonify({'success': True, 'project': serialize_project_full(doc)})


@app.route('/api/projects', methods=['POST'])
def create_project():
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    payload = request.get_json(silent=True) or {}
    cleaned, err = sanitize_project_payload(payload, require_data=True)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    pid = clean_project_id(payload.get('id'))
    ts = now_ms()

    if projects_col.find_one({'_id': pid}, {'_id': 1}):
        pid = generate_project_id()

    doc = {
        '_id': pid,
        'name': cleaned['name'],
        'data': cleaned['data'],
        'createdAt': int(payload.get('createdAt') or ts),
        'updatedAt': int(payload.get('updatedAt') or ts),
    }

    try:
        projects_col.insert_one(doc)
    except PyMongoError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500

    app.logger.info("POST /api/projects -> created id=%s name=%s summary=%s", pid, cleaned['name'], summarize_project_data(cleaned['data']))
    return jsonify({'success': True, 'project': serialize_project_full(doc)}), 201


@app.route('/api/projects/<project_id>', methods=['PUT'])
def update_project(project_id):
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    payload = request.get_json(silent=True) or {}
    cleaned, err = sanitize_project_payload(payload, require_data=False)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    update = {'updatedAt': now_ms()}
    if cleaned['name'] is not None:
        update['name'] = cleaned['name']
    if cleaned['data'] is not None:
        update['data'] = cleaned['data']

    if len(update) == 1:
        return jsonify({'success': False, 'error': 'No fields to update'}), 400

    pid = clean_project_id(project_id)
    before = projects_col.find_one({'_id': pid}, {'name': 1, 'updatedAt': 1})
    if not before:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    base_updated_at = payload.get('baseUpdatedAt')
    if base_updated_at is not None:
        try:
            expected = int(base_updated_at)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'error': 'baseUpdatedAt must be a valid integer timestamp.'}), 400
        current_updated_at = int(before.get('updatedAt') or 0)
        if expected != current_updated_at:
            latest_doc = projects_col.find_one({'_id': pid})
            return jsonify({
                'success': False,
                'error': 'Conflict: this project was updated in another tab/session. Reloaded latest server copy.',
                'project': serialize_project_full(latest_doc) if latest_doc else None
            }), 409

    summary = summarize_project_data(cleaned['data']) if cleaned['data'] is not None else {'valid': False, 'note': 'no data payload'}

    result = projects_col.update_one({'_id': pid}, {'$set': update})

    doc = projects_col.find_one({'_id': pid})
    app.logger.info(
        "PUT /api/projects/%s -> matched=%s modified=%s beforeUpdatedAt=%s afterUpdatedAt=%s name=%s payloadSummary=%s",
        pid,
        result.matched_count,
        result.modified_count,
        before.get('updatedAt') if before else None,
        doc.get('updatedAt') if doc else None,
        update.get('name'),
        summary
    )
    return jsonify({'success': True, 'project': serialize_project_full(doc)})


@app.route('/api/projects/<project_id>', methods=['DELETE'])
def delete_project(project_id):
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    docs = list(projects_col.find({}, {'_id': 1}))
    if len(docs) <= 1:
        return jsonify({'success': False, 'error': 'Cannot delete the only project.'}), 400

    pid = clean_project_id(project_id)
    result = projects_col.delete_one({'_id': pid})
    if result.deleted_count == 0:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    app.logger.info("DELETE /api/projects/%s -> deleted=%s", pid, result.deleted_count)
    return jsonify({'success': True})


@app.route('/api/projects/export', methods=['GET'])
def export_projects():
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    docs = list(projects_col.find({}).sort('updatedAt', DESCENDING))
    app.logger.info("GET /api/projects/export -> %s projects", len(docs))
    return jsonify({'success': True, 'projects': [serialize_project_full(d) for d in docs]})


@app.route('/api/projects/import', methods=['POST'])
def import_projects():
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    payload = request.get_json(silent=True) or {}
    mode = str(payload.get('mode') or 'merge').lower()
    raw_projects = payload.get('projects')

    if not isinstance(raw_projects, list):
        return jsonify({'success': False, 'error': 'projects must be an array.'}), 400
    if len(raw_projects) == 0:
        return jsonify({'success': False, 'error': 'No projects provided.'}), 400
    if len(raw_projects) > 500:
        return jsonify({'success': False, 'error': 'Too many projects in one import.'}), 400

    ts = now_ms()
    ops = []
    imported = 0

    for item in raw_projects:
        if not isinstance(item, dict):
            continue
        data = item.get('data')
        if not isinstance(data, dict):
            continue

        pid = clean_project_id(item.get('id'))
        name = clean_name(item.get('name'))
        created_at = int(item.get('createdAt') or ts)
        updated_at = int(item.get('updatedAt') or ts)

        ops.append(UpdateOne(
            {'_id': pid},
            {'$set': {
                'name': name,
                'data': data,
                'updatedAt': updated_at,
            }, '$setOnInsert': {
                'createdAt': created_at,
            }},
            upsert=True,
        ))
        imported += 1

    if imported == 0:
        return jsonify({'success': False, 'error': 'No valid projects in import payload.'}), 400

    try:
        if mode == 'replace':
            projects_col.delete_many({})
        projects_col.bulk_write(ops, ordered=False)
    except PyMongoError as exc:
        return jsonify({'success': False, 'error': str(exc)}), 500

    app.logger.info("POST /api/projects/import -> imported=%s mode=%s", imported, mode)
    return jsonify({'success': True, 'imported': imported, 'mode': mode})


@app.route('/api/projects/<project_id>/excel', methods=['GET'])
def export_project_excel(project_id):
    ready, error = ensure_db_ready()
    if not ready:
        return jsonify({'success': False, 'error': error}), 503

    pid = clean_project_id(project_id)
    doc = projects_col.find_one({'_id': pid})
    if not doc:
        return jsonify({'success': False, 'error': 'Project not found'}), 404

    name = clean_name(doc.get('name') or 'Project')
    data = doc.get('data') if isinstance(doc.get('data'), dict) else {}

    wb = build_excel_workbook(name, data)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    stamp = datetime.utcnow().strftime('%Y%m%d-%H%M%S')
    filename = f"{safe_filename(name)}-model-{stamp}.xlsx"
    app.logger.info("GET /api/projects/%s/excel -> generated workbook", pid)
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/api/save', methods=['POST'])
def save_deprecated():
    """Deprecated endpoint retained for backward compatibility."""
    return jsonify({
        'success': True,
        'deprecated': True,
        'message': 'Deprecated. Use /api/projects endpoints for persistence.'
    })


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
